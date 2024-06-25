import openpyxl   
from selenium import webdriver  #Abrir o nevegador
from selenium.webdriver.common.by import By  #Possibilita encontrar funções no site para interagir com ele.
from time import sleep

planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')  #Abrir o excel
pagina_clientes = planilha_clientes['Sheet1']  #Sheet1 é o nome da planilha

driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')

for linha in pagina_clientes.iter_rows(min_row=2,values_only=True):
   nome, valor, cpf, vencimento = linha   #Passando os parâmetros para extrair as informações (nome,valor,cpf,vencimento)

   sleep(5)
   campo_pesquisa = driver.find_element(By.XPATH,"//input[@id='cpfInput']")  #Tag do html
   sleep(1)
   campo_pesquisa.clear()
   campo_pesquisa.send_keys(cpf)  #Send_keys função do selenium que permite escrever em um campo dentro do site
   sleep(1)
   botao_pesquisar = driver.find_element(By.XPATH,"//button[@class='btn btn-custom btn-lg btn-block mt-3']")  #Tag do html-botão
   sleep(1)
   botao_pesquisar.click() 
   sleep(4)
   status = driver.find_element(By.XPATH,"//span[@id='statusLabel']")
   if status.text == 'em dia':
      data_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentDate']") 
      metodo_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentMethod']")
      
      data_pagamento_limpo = data_pagamento.text.split()[3]
      metodo_pagamento_limpo = metodo_pagamento.text.split()[3]
      
      planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
      pagina_fechamento = planilha_fechamento['Sheet1']
   
      pagina_fechamento.append([nome, valor, cpf, vencimento,'em dia', 
      data_pagamento_limpo,metodo_pagamento_limpo])

      planilha_fechamento.save('planilha fechamento.xlsx')
   else:
     planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
     pagina_fechamento = planilha_fechamento['Sheet1']

     pagina_fechamento.append([nome, valor, cpf, vencimento,'pendente'])
     planilha_fechamento.save('planilha fechamento.xlsx')